from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def generar_pdf_boleta(alumno, matricula, grupo, especialidad, filas):


    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    estilos = getSampleStyleSheet()

    elementos = []

    titulo = Paragraph(
        "<b>BOLETA DE CALIFICACIONES</b>",
        estilos["Title"]
    )

    elementos.append(titulo)
    elementos.append(Spacer(1, 20))

    estilo_info = ParagraphStyle(
        "info",
        parent=estilos["Normal"],
        alignment=1
    )

    if especialidad:
        grupo_texto = f"{grupo} - {especialidad}"
    else:
        grupo_texto = grupo

    tabla_info = Table(
        [[
            Paragraph(f"<b>Alumno:</b> {alumno}", estilo_info),
            Paragraph(f"<b>Matricula:</b> {matricula}", estilo_info),
            Paragraph(f"<b>Grupo:</b> {grupo_texto}", estilo_info),
            Paragraph(f"<b>Fecha:</b> {date.today()}", estilo_info)
        ]],
        colWidths=[7*cm, 4*cm, 5*cm]
    )

    tabla_info.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))

    elementos.append(tabla_info)
    elementos.append(Spacer(1, 20))

    encabezado = [
        "Materia",
        "Parcial 1",
        "Parcial 2",
        "Parcial 3",
        "Final",
        "Tipo",
        "Periodo"
    ]

    datos_tabla = [encabezado] + filas

    tabla = Table(datos_tabla)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("ALIGN", (1,1), (-2,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    return pdf



def generar_pdf_kardex(alumno, escuela, datos_por_grupo, promedio_general):

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=1*cm,
        bottomMargin=2*cm
    )

    estilos = getSampleStyleSheet()

    estilo_centro = ParagraphStyle(
        "centro",
        parent=estilos["Normal"],
        alignment=1
    )

    estilo_escuela = ParagraphStyle(
        "escuela",
        parent=estilos["Normal"],
        alignment=1,
        fontSize=10
    )

    elementos = []

    # =========================
    # ENCABEZADO ESCUELA
    # =========================

    elementos.append(
        Paragraph(f"<b>{escuela['nombre']}</b>", estilo_centro)
    )

    elementos.append(
        Paragraph(f"RFC: {escuela['rfc']}", estilo_escuela)
    )

    elementos.append(
        Paragraph(f"{escuela['direccion']}", estilo_escuela)
    )

    elementos.append(Spacer(1,20))

    # =========================
    # TITULO
    # =========================

    titulo = Paragraph(
        "<b>KARDEX DEL ALUMNO</b>",
        estilos["Title"]
    )

    elementos.append(titulo)
    elementos.append(Spacer(1,20))

    info = Paragraph(
        f"<b>Alumno:</b> {alumno} &nbsp;&nbsp;&nbsp;&nbsp; <b>Fecha:</b> {date.today()}",
        estilo_centro
    )

    elementos.append(info)
    elementos.append(Spacer(1,25))

    encabezado = [
        "Clave",
        "Materia",
        "Final",
        "Tipo",
        "Periodo"
    ]

    for grupo, datos in datos_por_grupo.items():

        titulo_grupo = grupo

        if datos["especialidad"]:
            titulo_grupo += f"   Especialidad: {datos['especialidad']}"

        elementos.append(
            Paragraph(f"<b>{titulo_grupo}</b>", estilos["Heading3"])
        )

        tabla_datos = [encabezado] + datos["filas"]

        tabla = Table(
            tabla_datos,
            colWidths=[
                1.7*cm,
                8*cm,
                1.3*cm,
                1.2*cm,
                3.8*cm
            ],
            repeatRows=1
        )

        tabla.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
            ("GRID",(0,0),(-1,-1),0.5,colors.black),
            ("ALIGN",(2,1),(-2,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ]))

        elementos.append(tabla)

        elementos.append(Spacer(1,5))

        elementos.append(
            Paragraph(
                f"<b>Promedio:</b> {datos['promedio']:.2f}",
                estilos["Normal"]
            )
        )

        elementos.append(Spacer(1,20))

    elementos.append(
        Paragraph(
            f"<b>Promedio general:</b> {promedio_general:.2f}",
            estilos["Heading2"]
        )
    )

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    return pdf



def generar_pdf_reprobados(escuela, grado, grupo, filas):

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    estilos = getSampleStyleSheet()

    estilo_centro = ParagraphStyle(
        "centro",
        parent=estilos["Normal"],
        alignment=1
    )

    elementos = []

    # =========================
    # ENCABEZADO ESCUELA
    # =========================

    elementos.append(
        Paragraph(f"<b>{escuela['nombre']}</b>", estilo_centro)
    )

    elementos.append(
        Paragraph(f"RFC: {escuela['rfc']}", estilo_centro)
    )

    elementos.append(
        Paragraph(f"{escuela['direccion']}", estilo_centro)
    )

    elementos.append(Spacer(1,20))

    # =========================
    # TITULO
    # =========================

    elementos.append(
        Paragraph("<b>REPORTE DE ALUMNOS REPROBADOS</b>", estilos["Title"])
    )

    elementos.append(Spacer(1,15))

    info = Paragraph(
        f"<b>Grado:</b> {grado} &nbsp;&nbsp;&nbsp;&nbsp; "
        f"<b>Grupo:</b> {grupo} &nbsp;&nbsp;&nbsp;&nbsp; "
        f"<b>Fecha:</b> {date.today()}",
        estilo_centro
    )

    elementos.append(info)

    elementos.append(Spacer(1,20))

    encabezado = [
        "No",
        "Alumno",
        "Materia",
        "Final"
    ]

    tabla_datos = [encabezado] + filas

    tabla = Table(
        tabla_datos,
        colWidths=[1*cm,9*cm,7.5*cm,1.2*cm],
        repeatRows=1
    )

    tabla.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
        ("GRID",(0,0),(-1,-1),0.5,colors.black),
        ("ALIGN",(3,1),(3,-1),"CENTER"),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    return pdf



def generar_excel_boleta_materia(escuela, grupo, materia, alumnos, periodo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Boleta"

    # =========================
    # ESTILOS
    # =========================

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # =========================
    # ENCABEZADO ESCUELA
    # =========================

    ws.merge_cells("A1:F1")
    ws["A1"] = escuela["nombre"]
    ws["A1"].font = bold
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"RFC: {escuela['rfc']}"
    ws["A2"].alignment = center

    ws.merge_cells("A3:F3")
    ws["A3"] = escuela["direccion"]
    ws["A3"].alignment = center

    # =========================
    # TITULO
    # =========================

    ws.merge_cells("A5:F5")
    ws["A5"] = "CONCENTRADO DE CALIFICACIONES"
    ws["A5"].font = bold
    ws["A5"].alignment = center

    # =========================
    # DATOS
    # =========================

    ws["A7"] = "Maestro:"
    ws["A7"].font = bold
    ws.merge_cells("A8:B8")
    
    ws["D7"] = "Grupo:"
    ws["D7"].font = bold
    ws.merge_cells("E7:F7")
    ws["E7"] = grupo
    ws["E7"].alignment = center

    ws["D8"] = "Periodo:"
    ws["D8"].font = bold
    ws.merge_cells("E8:F8")
    ws["E8"] = periodo
    ws["E8"].alignment = center


    ws.merge_cells("A10:F10")
    ws["A10"] = materia
    ws["A10"].font = bold
    ws["A10"].alignment = center

    # =========================
    # ENCABEZADO TABLA
    # =========================

    headers = ["Mat.", "Alumno", "Parcial 1", "Parcial 2", "Parcial 3", "Final"]

    row = 12

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = h
        cell.font = bold
        cell.border = border
        cell.alignment = center

    # =========================
    # FILAS (VACÍAS PARA LLENAR)
    # =========================

    row += 1

    for fila in alumnos:
        matricula, alumno, p1, p2, p3, final = fila

        ws.cell(row=row, column=1).value = matricula
        ws.cell(row=row, column=2).value = alumno
        ws.cell(row=row, column=3).value = p1
        ws.cell(row=row, column=4).value = p2
        ws.cell(row=row, column=5).value = p3
        ws.cell(row=row, column=6).value = final

        for col in range(1,7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = center

        row += 1

    # =========================
    # FIRMA
    # =========================

    row += 2

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws.cell(row=row, column=3).value = "FIRMA DEL ASESOR"
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).font = bold

    thin = Side(style="thin")

    # Dimensiones del cuadro
    inicio_row = row + 1
    fin_row = row + 3
    inicio_col = 3
    fin_col = 5

    none = Side(style=None)

    for r in range(inicio_row, fin_row + 1):
        for c in range(inicio_col, fin_col + 1):

            left = thin if c == inicio_col else none
            right = thin if c == fin_col else none
            top = thin if r == inicio_row else none
            bottom = thin if r == fin_row else none

            ws.cell(row=r, column=c).border = Border(
                left=left,
                right=right,
                top=top,
                bottom=bottom
            )

    # =========================
    # AJUSTES
    # =========================

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 43
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 11

    buffer = BytesIO()
    wb.save(buffer)

    excel_bytes = buffer.getvalue()
    buffer.close()

    return excel_bytes



def generar_excel_plantilla(escuela, grupo, materia, alumnos, periodo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Boleta"

    # =========================
    # ESTILOS
    # =========================

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # =========================
    # ENCABEZADO ESCUELA
    # =========================

    ws.merge_cells("A1:F1")
    ws["A1"] = escuela["nombre"]
    ws["A1"].font = bold
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"RFC: {escuela['rfc']}"
    ws["A2"].alignment = center

    ws.merge_cells("A3:F3")
    ws["A3"] = escuela["direccion"]
    ws["A3"].alignment = center

    # =========================
    # TITULO
    # =========================

    ws.merge_cells("A5:F5")
    ws["A5"] = "CONCENTRADO DE CALIFICACIONES"
    ws["A5"].font = bold
    ws["A5"].alignment = center

    # =========================
    # DATOS
    # =========================

    ws["A7"] = "Maestro:"
    ws["A7"].font = bold
    ws.merge_cells("A8:B8")
    
    ws["D7"] = "Grupo:"
    ws["D7"].font = bold
    ws.merge_cells("E7:F7")
    ws["E7"] = grupo
    ws["E7"].alignment = center

    ws["D8"] = "Periodo:"
    ws["D8"].font = bold
    ws.merge_cells("E8:F8")
    ws["E8"] = periodo
    ws["E8"].alignment = center


    ws.merge_cells("A10:F10")
    ws["A10"] = materia
    ws["A10"].font = bold
    ws["A10"].alignment = center

    # =========================
    # ENCABEZADO TABLA
    # =========================

    headers = ["Mat.", "Alumno", "Parcial 1", "Parcial 2", "Parcial 3", "Final"]

    row = 12

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = h
        cell.font = bold
        cell.border = border
        cell.alignment = center

    # =========================
    # FILAS (VACÍAS PARA LLENAR)
    # =========================

    row += 1

    for fila in alumnos:
        matricula, alumno = fila
        ws.cell(row=row, column=1).value = matricula
        ws.cell(row=row, column=2).value = alumno

        for col in range(1,7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = center

        row += 1

    # =========================
    # FIRMA
    # =========================

    row += 2

    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws.cell(row=row, column=3).value = "FIRMA DEL ASESOR"
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).font = bold

    thin = Side(style="thin")

    # Dimensiones del cuadro
    inicio_row = row + 1
    fin_row = row + 3
    inicio_col = 3
    fin_col = 5

    none = Side(style=None)

    for r in range(inicio_row, fin_row + 1):
        for c in range(inicio_col, fin_col + 1):

            left = thin if c == inicio_col else none
            right = thin if c == fin_col else none
            top = thin if r == inicio_row else none
            bottom = thin if r == fin_row else none

            ws.cell(row=r, column=c).border = Border(
                left=left,
                right=right,
                top=top,
                bottom=bottom
            )

    # =========================
    # AJUSTES
    # =========================

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 43
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 11

    buffer = BytesIO()
    wb.save(buffer)

    excel_bytes = buffer.getvalue()
    buffer.close()

    return excel_bytes









